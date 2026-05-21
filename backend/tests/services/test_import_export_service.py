import base64
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app.import_export import service as import_export_service
from app.orders.models import SalesOrderCreate, SalesOrderItemCreate
from app.orders import service as orders_service
from app.product.models import Product

PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
)


def _write_image(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(PNG_BYTES)


def test_export_excel_embeds_product_images(session, tmp_path, monkeypatch):
    image_root = tmp_path / "img"
    monkeypatch.setattr(import_export_service, "IMG_DIR", image_root)
    _write_image(image_root / "sample.png")
    session.add(Product(name="羽毛球", image="img/sample.png"))
    session.commit()

    result = import_export_service.export_excel(session, str(tmp_path / "export"))

    workbook = load_workbook(result["path"])
    sheet = workbook["产品"]
    assert sheet["B2"].value is None
    assert len(sheet._images) == 1


def test_export_order_excel_exports_single_order_with_image(session, tmp_path, monkeypatch):
    image_root = tmp_path / "img"
    monkeypatch.setattr(import_export_service, "IMG_DIR", image_root)
    _write_image(image_root / "order-product.png")
    order = orders_service.create_order(
        session,
        SalesOrderCreate(
            customer_name="张三",
            customer_phone="13800000000",
            delivery_address="上海",
            sales_date=date(2026, 4, 20),
            delivery_date=date(2026, 4, 22),
            payment_terms="月结",
            notes="备注",
            items=[
                SalesOrderItemCreate(
                    product_name="羽毛球",
                    total_boxes=2,
                    per_box_qty=12,
                    unit_price=Decimal("10.50"),
                    image="img/order-product.png",
                )
            ],
        ),
    )

    result = import_export_service.export_order_excel(session, order.id, str(tmp_path / "export"))

    workbook = load_workbook(result["path"])
    sheet = workbook["销售单"]
    assert sheet["B3"].value == "张三"
    assert sheet["B7"].value == "羽毛球"
    assert len(sheet._images) == 1
