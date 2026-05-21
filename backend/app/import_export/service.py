from __future__ import annotations

import base64
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, SQLModel, select

from app.config import IMG_DIR
from app.customer.models import Customer
from app.orders.service import get_order
from app.product.images import store_product_image
from app.product.models import Product
from app.purchases.models import PurchaseOrder
from app.sales.models import SalesRecord
from app.supplier.models import Supplier
from app.tasks.models import Task


class ImportExportError(Exception):
    pass


@dataclass(frozen=True)
class FieldSpec:
    name: str
    label: str
    required: bool = False
    kind: str = "str"


@dataclass(frozen=True)
class TableSpec:
    key: str
    sheet_name: str
    model: type[SQLModel]
    fields: tuple[FieldSpec, ...]
    unique_field: str | None = None


TABLE_SPECS = (
    TableSpec(
        "customers",
        "客户",
        Customer,
        (
            FieldSpec("name", "客户名称", True),
            FieldSpec("phone", "电话"),
            FieldSpec("address", "地址"),
            FieldSpec("notes", "备注"),
        ),
        "name",
    ),
    TableSpec(
        "suppliers",
        "厂家",
        Supplier,
        (
            FieldSpec("name", "厂家名称", True),
            FieldSpec("phone", "电话"),
            FieldSpec("address", "地址"),
            FieldSpec("notes", "备注"),
        ),
        "name",
    ),
    TableSpec(
        "products",
        "产品",
        Product,
        (
            FieldSpec("name", "产品名称", True),
            FieldSpec("image", "图片"),
            FieldSpec("per_box_qty", "装箱数", kind="int"),
            FieldSpec("box_spec", "箱规"),
            FieldSpec("volume", "体积", kind="decimal"),
            FieldSpec("purchase_price", "进货价格", kind="decimal"),
            FieldSpec("stock_qty", "库存数量", kind="int"),
            FieldSpec("notes", "备注"),
        ),
        "name",
    ),
    TableSpec(
        "sales",
        "销售",
        SalesRecord,
        (
            FieldSpec("sale_time", "销售时间", True, "date"),
            FieldSpec("customer_name", "客户名称", True),
            FieldSpec("product", "产品", True),
            FieldSpec("amount", "销售金额", True, "decimal"),
            FieldSpec("delivery_time", "送货时间", kind="date"),
            FieldSpec("collection_time", "收款时间", kind="date"),
            FieldSpec("is_settled", "是否结清", kind="bool"),
            FieldSpec("payment_method", "交易方式"),
            FieldSpec("cost", "成本", True, "decimal"),
            FieldSpec("notes", "备注"),
        ),
    ),
    TableSpec(
        "purchases",
        "采购",
        PurchaseOrder,
        (
            FieldSpec("purchase_time", "采购时间", True, "date"),
            FieldSpec("supplier_name", "厂家名称", True),
            FieldSpec("product_name", "产品名称", True),
            FieldSpec("box_count", "箱数", True, "int"),
            FieldSpec("per_box_qty", "装箱数", True, "int"),
            FieldSpec("unit_price", "单价", True, "decimal"),
            FieldSpec("paid_amount", "已付金额", kind="decimal"),
            FieldSpec("notes", "备注"),
        ),
    ),
    TableSpec(
        "tasks",
        "任务",
        Task,
        (
            FieldSpec("title", "标题", True),
            FieldSpec("description", "描述"),
            FieldSpec("category", "分类"),
            FieldSpec("priority", "优先级"),
            FieldSpec("status", "状态"),
            FieldSpec("due_date", "到期日期", kind="date"),
            FieldSpec("related_type", "关联类型"),
            FieldSpec("related_id", "关联ID", kind="nullable_int"),
            FieldSpec("notes", "备注"),
        ),
    ),
)


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bool):
        return "是" if value else "否"
    return value


def _rows_for_export(session: Session, spec: TableSpec) -> list[dict[str, Any]]:
    rows = session.exec(select(spec.model)).all()
    return [
        {field.label: _cell_value(getattr(row, field.name, "")) for field in spec.fields}
        for row in rows
    ]


def _safe_filename(name: str, fallback: str) -> str:
    safe = "".join("-" if char in '\\/:*?"<>|\r\n\t' else char for char in name.strip()).strip(" .-")
    return safe or fallback


def _resolve_local_image(image: str) -> Path | BytesIO | None:
    if not image:
        return None
    if image.startswith("data:"):
        try:
            _, encoded = image.split(",", 1)
            return BytesIO(base64.b64decode(encoded))
        except ValueError:
            return None

    if image.startswith("http"):
        return None

    normalized = image.lstrip("/")
    if normalized.startswith("img/"):
        path = IMG_DIR / normalized.removeprefix("img/")
    else:
        path = Path(image).expanduser()
    return path if path.exists() else None


def _fit_excel_image(image: ExcelImage, max_width: int = 92, max_height: int = 72) -> None:
    if image.width <= 0 or image.height <= 0:
        return
    scale = min(max_width / image.width, max_height / image.height, 1)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)


def _embed_image(ws, cell: str, image_value: str) -> bool:
    image_source = _resolve_local_image(image_value)
    if not image_source:
        return False
    try:
        image = ExcelImage(image_source)
    except Exception:
        return False
    _fit_excel_image(image)
    ws.add_image(image, cell)
    return True


def _write_table_sheet(workbook: Workbook, session: Session, spec: TableSpec) -> None:
    ws = workbook.create_sheet(spec.sheet_name)
    headers = [field.label for field in spec.fields]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAF2F8")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = session.exec(select(spec.model)).all()
    if not rows:
        return

    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(spec.fields, start=1):
            cell = ws.cell(row=row_index, column=column_index)
            value = getattr(row, field.name, "")
            if field.name == "image" and value:
                embedded = _embed_image(ws, cell.coordinate, str(value))
                if embedded:
                    ws.row_dimensions[row_index].height = 58
                    ws.column_dimensions[get_column_letter(column_index)].width = 15
                continue
            cell.value = _cell_value(value)

    for column_index, field in enumerate(spec.fields, start=1):
        column_letter = get_column_letter(column_index)
        current_width = ws.column_dimensions[column_letter].width or 10
        if field.name == "image":
            ws.column_dimensions[column_letter].width = max(current_width, 15)
        else:
            ws.column_dimensions[column_letter].width = max(current_width, min(max(len(field.label) + 4, 12), 24))


def export_excel(session: Session, target_dir: str) -> dict:
    if not target_dir.strip():
        raise ImportExportError("请先填写导出目录")

    target_dir = Path(target_dir).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"sis-book-export-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

    workbook = Workbook()
    workbook.remove(workbook.active)
    for spec in TABLE_SPECS:
        _write_table_sheet(workbook, session, spec)
    workbook.save(target_path)

    return {"cancelled": False, "path": str(target_path)}


def export_order_excel(session: Session, order_id: int, target_dir: str) -> dict:
    if not target_dir.strip():
        raise ImportExportError("请先填写导出目录")

    order = get_order(session, order_id)
    if not order:
        raise ImportExportError("销售单不存在")

    target_dir = Path(target_dir).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)
    order_number = order.order_number or f"sales-order-{order.id}"
    target_path = target_dir / f"{_safe_filename(order_number, '销售单')}.xlsx"

    workbook = Workbook()
    ws = workbook.active
    ws.title = "销售单"
    ws.merge_cells("A1:K1")
    ws["A1"] = "暮橙体育销售单"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["客户", order.customer_name, "电话", order.customer_phone, "销售日期", order.sales_date.isoformat()])
    ws.append(["送货地址", order.delivery_address, "送货日期", order.delivery_date.isoformat() if order.delivery_date else "", "付款方式", order.payment_terms])
    ws.append([])

    headers = ["编号", "产品", "颜色", "图片", "总箱数", "每箱数量", "总数量", "单价", "金额", "外箱尺寸", "备注"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D6EAF8")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_boxes = 0
    total_qty = 0
    total_amount = Decimal("0")
    for index, item in enumerate(order.items, start=1):
        total_qty_for_item = item.total_boxes * item.per_box_qty
        subtotal = Decimal(item.total_boxes) * Decimal(item.per_box_qty) * Decimal(str(item.unit_price))
        total_boxes += item.total_boxes
        total_qty += total_qty_for_item
        total_amount += subtotal
        ws.append([
            index,
            item.product_name,
            item.color_spec,
            "",
            item.total_boxes,
            item.per_box_qty,
            total_qty_for_item,
            float(item.unit_price),
            float(subtotal),
            item.box_size,
            item.notes,
        ])
        row_index = ws.max_row
        if item.image and _embed_image(ws, f"D{row_index}", item.image):
            ws.row_dimensions[row_index].height = 58

    ws.append(["合计", "", "", "", total_boxes, "", total_qty, "", float(total_amount), "", ""])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAF2F8")

    if order.notes:
        ws.append([])
        ws.append(["备注", order.notes])

    widths = [8, 22, 14, 15, 10, 12, 10, 10, 12, 14, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(target_path)
    return {"cancelled": False, "path": str(target_path)}


def _read_workbook(payload: bytes) -> dict[str, pd.DataFrame]:
    if not payload:
        raise ImportExportError("请选择 Excel 文件")
    try:
        return pd.read_excel(BytesIO(payload), sheet_name=None, dtype=str, keep_default_na=False)
    except Exception as exc:
        raise ImportExportError("Excel 文件读取失败") from exc


def _normalized(text: str) -> str:
    return text.strip().lower().replace(" ", "").replace("_", "")


def _field_lookup(spec: TableSpec) -> dict[str, FieldSpec]:
    lookup: dict[str, FieldSpec] = {}
    for field in spec.fields:
        lookup[_normalized(field.name)] = field
        lookup[_normalized(field.label)] = field
    return lookup


def _match_spec(sheet_name: str, columns: list[str]) -> TableSpec | None:
    normalized_sheet = _normalized(sheet_name)
    for spec in TABLE_SPECS:
        if normalized_sheet in {_normalized(spec.key), _normalized(spec.sheet_name)}:
            return spec

    best_spec: TableSpec | None = None
    best_score = 0
    for spec in TABLE_SPECS:
        lookup = _field_lookup(spec)
        score = sum(1 for column in columns if _normalized(str(column)) in lookup)
        if score > best_score:
            best_score = score
            best_spec = spec
    return best_spec if best_score >= 2 else None


def _column_mapping(spec: TableSpec, columns: list[str]) -> dict[str, FieldSpec]:
    lookup = _field_lookup(spec)
    return {
        column: lookup[_normalized(str(column))]
        for column in columns
        if _normalized(str(column)) in lookup
    }


def preview_excel(payload: bytes) -> dict:
    workbook = _read_workbook(payload)
    sheets = []
    for sheet_name, frame in workbook.items():
        columns = [str(column) for column in frame.columns]
        spec = _match_spec(sheet_name, columns)
        mapping = _column_mapping(spec, columns) if spec else {}
        mapped_fields = {field.name for field in mapping.values()}
        warnings = []
        if spec:
            missing_required = [
                field.label for field in spec.fields if field.required and field.name not in mapped_fields
            ]
            if missing_required:
                warnings.append(f"缺少必填列：{'、'.join(missing_required)}")
        else:
            warnings.append("未匹配到可导入的数据表")

        sheets.append(
            {
                "name": sheet_name,
                "matched_table": spec.key if spec else "",
                "matched_label": spec.sheet_name if spec else "",
                "total_rows": len(frame.index),
                "columns": columns,
                "mappings": [
                    {
                        "column": column,
                        "field": field.name,
                        "label": field.label,
                        "required": field.required,
                    }
                    for column, field in mapping.items()
                ],
                "rows": frame.head(10).to_dict(orient="records"),
                "warnings": warnings,
            }
        )
    return {"sheets": sheets}


def _empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _to_date(value: Any) -> date | None:
    if _empty(value):
        return None
    parsed = pd.to_datetime(value, errors="raise")
    return parsed.date()


def _to_bool(value: Any) -> bool:
    if _empty(value):
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是", "已结清", "结清"}


def _coerce_value(field: FieldSpec, value: Any) -> Any:
    if field.kind == "date":
        return _to_date(value)
    if field.kind == "bool":
        return _to_bool(value)
    if field.kind == "int":
        return 0 if _empty(value) else int(Decimal(str(value)))
    if field.kind == "nullable_int":
        return None if _empty(value) else int(Decimal(str(value)))
    if field.kind == "decimal":
        return Decimal("0") if _empty(value) else Decimal(str(value))
    return "" if value is None else str(value).strip()


def _payload_from_row(spec: TableSpec, mapping: dict[str, FieldSpec], row: dict) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for column, field in mapping.items():
        payload[field.name] = _coerce_value(field, row.get(column, ""))

    missing_required = [
        field.label
        for field in spec.fields
        if field.required and _empty(payload.get(field.name))
    ]
    if missing_required:
        raise ImportExportError(f"{spec.sheet_name} 缺少必填值：{'、'.join(missing_required)}")

    if spec.key == "products" and payload.get("image"):
        payload["image"] = store_product_image(payload["image"], payload["name"])
    return payload


def _upsert_or_create(session: Session, spec: TableSpec, payload: dict[str, Any]) -> str:
    if spec.unique_field:
        unique_value = payload.get(spec.unique_field)
        column = getattr(spec.model, spec.unique_field)
        entity = session.exec(select(spec.model).where(column == unique_value)).first()
        if entity:
            for key, value in payload.items():
                setattr(entity, key, value)
            if hasattr(entity, "updated_at"):
                setattr(entity, "updated_at", datetime.now())
            session.add(entity)
            return "updated"

    session.add(spec.model(**payload))
    return "created"


def import_excel(session: Session, payload: bytes) -> dict:
    workbook = _read_workbook(payload)
    summary = []
    for sheet_name, frame in workbook.items():
        columns = [str(column) for column in frame.columns]
        spec = _match_spec(sheet_name, columns)
        if not spec:
            continue

        mapping = _column_mapping(spec, columns)
        created = 0
        updated = 0
        for row in frame.to_dict(orient="records"):
            if all(_empty(value) for value in row.values()):
                continue
            try:
                payload = _payload_from_row(spec, mapping, row)
            except Exception as exc:
                raise ImportExportError(f"{spec.sheet_name} 第 {created + updated + 1} 行导入失败：{exc}") from exc
            action = _upsert_or_create(session, spec, payload)
            if action == "updated":
                updated += 1
            else:
                created += 1

        summary.append(
            {
                "sheet": sheet_name,
                "table": spec.sheet_name,
                "created": created,
                "updated": updated,
            }
        )

    session.commit()
    return {"summary": summary}
